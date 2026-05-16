"""
tests.test_pp_v1_4_4_ct_study_io — v1.4.4 Sprint E/F.

Cobre o módulo ``app.postprocessor.ct_saturation_io`` (JSON loader
+ TXT/XLSX writers) replicando ``main_CT_Analysis.m`` §1-2 e
``generate_reports.m``.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest


CT_JSON_PATH = Path(
    "D:/000 - UFMG - DOUTORADO/MVP/LIB/LIB/CT_STUDY/ct_system_data.json"
)


# ===========================================================================
# Sprint E — JSON loader (ct_system_data.json)
# ===========================================================================


class TestCTSystemDataLoader:

    def test_module_loadable(self):
        from app.postprocessor import ct_saturation_io  # noqa: F401

    def test_load_returns_list_of_fault_contexts(self):
        from app.postprocessor.ct_saturation import CTFaultContext
        from app.postprocessor.ct_saturation_io import (
            load_ct_system_data,
        )
        if not CT_JSON_PATH.exists():
            pytest.skip("ct_system_data.json não encontrado (LIB/CT_STUDY)")
        ctxs = load_ct_system_data(CT_JSON_PATH)
        assert isinstance(ctxs, list)
        assert len(ctxs) > 0
        for c in ctxs[:3]:
            assert isinstance(c, CTFaultContext)

    def test_first_entry_is_28kA_bus(self):
        """json[0] = 00-MF-32/42 com Icc=28.4 kA, X/R=13.02."""
        from app.postprocessor.ct_saturation_io import load_ct_system_data
        if not CT_JSON_PATH.exists():
            pytest.skip("ct_system_data.json não encontrado")
        ctxs = load_ct_system_data(CT_JSON_PATH)
        first = ctxs[0]
        assert "00-MF" in first.bus_id or "MF-32" in first.bus_id
        assert abs(first.I_fault_rms_kA - 28.4) < 0.01
        assert abs(first.X_over_R - 13.02) < 0.05

    def test_summary_table_includes_tag_and_ratio(self):
        """list_cts_table retorna [{'CT_TAG', 'CT_Ratio', 'I_fault_rms_kA', ...}]."""
        from app.postprocessor.ct_saturation_io import list_cts_table
        if not CT_JSON_PATH.exists():
            pytest.skip("ct_system_data.json não encontrado")
        table = list_cts_table(CT_JSON_PATH)
        assert len(table) > 10
        first = table[0]
        for k in ("index", "CT_TAG", "CT_Ratio", "I_fault_rms_kA", "X_over_R"):
            assert k in first


# ===========================================================================
# Sprint F — Report writers (TXT + XLSX)
# ===========================================================================


@pytest.fixture
def sample_report():
    """Report didático para testes de IO."""
    from app.postprocessor.ct_saturation import (
        CTFaultContext, CTSpec, analyze_full,
        typical_magnetization_curve,
    )
    fault = CTFaultContext(
        bus_id="BUS-TEST", rated_voltage_kV=13.8,
        I_fault_rms_kA=20.0, X_over_R=10.0,
    )
    ct = CTSpec(
        tag="CT-IO-1", tipo_TC="ANSI_C",
        I_pri_nom_tap=2000, I_pri_nom_full=2000, I_sn=5,
        CT_Rating_C=400.0, R_s=0.5, R_b=0.3,
        curve=typical_magnetization_curve(400.0),
        Kr=0.1,
    )
    return analyze_full(ct, fault)


class TestReportWriterTxt:

    def test_writes_file(self, sample_report, tmp_path):
        from app.postprocessor.ct_saturation_io import write_report_txt
        out = tmp_path / "Relatorio_test.txt"
        write_report_txt(sample_report, out)
        assert out.exists()
        assert out.stat().st_size > 100

    def test_contains_tag_and_status(self, sample_report, tmp_path):
        from app.postprocessor.ct_saturation_io import write_report_txt
        out = tmp_path / "Rep.txt"
        write_report_txt(sample_report, out)
        content = out.read_text(encoding="utf-8")
        assert "CT-IO-1" in content
        assert "BUS-TEST" in content
        assert "RESULTADO FINAL" in content

    def test_contains_three_sections(self, sample_report, tmp_path):
        """TXT deve ter 3 seções (Especificação, Sistema, Análise Dinâmica)."""
        from app.postprocessor.ct_saturation_io import write_report_txt
        out = tmp_path / "Rep.txt"
        write_report_txt(sample_report, out)
        content = out.read_text(encoding="utf-8")
        assert "ESPECIFICA" in content   # Especificação
        assert "SISTEMA" in content
        assert "DIN" in content          # Dinâmica/Análise Dinâmica


class TestReportWriterXlsx:

    def test_module_imports_openpyxl(self):
        """Pré-requisito: openpyxl está disponível."""
        pytest.importorskip("openpyxl")

    def test_writes_xlsx(self, sample_report, tmp_path):
        pytest.importorskip("openpyxl")
        from app.postprocessor.ct_saturation_io import write_report_xlsx
        out = tmp_path / "Datasheet_test.xlsx"
        write_report_xlsx(sample_report, out)
        assert out.exists()
        assert out.stat().st_size > 100

    def test_xlsx_has_8_rows(self, sample_report, tmp_path):
        """Tabela 2 colunas × 8 linhas (paridade generate_reports.m)."""
        pytest.importorskip("openpyxl")
        from app.postprocessor.ct_saturation_io import write_report_xlsx
        from openpyxl import load_workbook
        out = tmp_path / "Datasheet_test.xlsx"
        write_report_xlsx(sample_report, out)
        wb = load_workbook(out)
        ws = wb.active
        # Header + 8 rows
        assert ws.max_row >= 8

    def test_xlsx_contains_tag(self, sample_report, tmp_path):
        pytest.importorskip("openpyxl")
        from app.postprocessor.ct_saturation_io import write_report_xlsx
        from openpyxl import load_workbook
        out = tmp_path / "Datasheet_test.xlsx"
        write_report_xlsx(sample_report, out)
        wb = load_workbook(out)
        ws = wb.active
        all_values = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    all_values.append(str(cell))
        joined = " ".join(all_values)
        assert "CT-IO-1" in joined


# ===========================================================================
# Sprint D — next_ansi_class helper (workflow refinamento)
# ===========================================================================


class TestNextAnsiClass:

    def test_progression_100_to_200(self):
        from app.postprocessor.ct_saturation_io import next_ansi_class
        assert next_ansi_class(100.0) == 200.0

    def test_progression_200_to_400(self):
        from app.postprocessor.ct_saturation_io import next_ansi_class
        assert next_ansi_class(200.0) == 400.0

    def test_progression_400_to_800(self):
        from app.postprocessor.ct_saturation_io import next_ansi_class
        assert next_ansi_class(400.0) == 800.0

    def test_above_800_returns_at_least_800(self):
        """Acima de C800 → custom (mantém ou sobe)."""
        from app.postprocessor.ct_saturation_io import next_ansi_class
        result = next_ansi_class(800.0)
        assert result >= 800.0

    def test_intermediate_value_rounds_up(self):
        """C150 → próxima padrão é C200."""
        from app.postprocessor.ct_saturation_io import next_ansi_class
        assert next_ansi_class(150.0) == 200.0
