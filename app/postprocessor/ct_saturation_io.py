"""
app.postprocessor.ct_saturation_io — v1.4.4 Sprint E/F + v1.4.5 fixes.

I/O helpers para o CT Saturation Study, replicando paridade com:

* ``main_CT_Analysis.m`` §1-2 (carregamento de listas JSON de TCs)
* ``generate_reports.m`` (TXT + XLSX)

Inclui também:

* ``next_ansi_class`` — helper do workflow de refinamento (Sprint D)
* ``CTSelection`` — bridge entre PpComponent CT do canvas e
  CTSpec/CTFaultContext (v1.4.5 — workflow gateado por seleção)

Notas sobre formato JSON
-------------------------

O loader lê **listas de TCs em formato JSON** com schema neutro
(CT_TAG, Bus_in, V_LL_kV, I_fault_rms_kA, X_over_R, CT_Ratio).
Esse formato é **compatível com exports** de outras ferramentas
(SKM PTW Power*Tools, ETAP, EasyPower), mas o Olivas **não tem
integração ativa** com nenhuma — apenas lê o arquivo JSON estático.
"""

from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Union

from app.core.logging_config import get_logger
from app.postprocessor.ct_saturation import (
    CTFaultContext, CTSaturationReport,
)

log = get_logger(__name__)

PathLike = Union[str, Path]


# ---------------------------------------------------------------------------
# Sprint E — JSON loader (ct_system_data.json)
# ---------------------------------------------------------------------------


def load_ct_system_data(path: PathLike) -> List[CTFaultContext]:
    """
    Carrega lista JSON de TCs (formato compatível com exports de
    PTW/ETAP/EasyPower; sem integração ativa) e converte cada
    entrada em :class:`CTFaultContext`.

    Schema esperado (por TC):

    * ``CT_TAG`` (str)
    * ``Bus_in`` (str) — usado como bus_id (com fallback para CT_TAG)
    * ``V_LL_kV`` (float)
    * ``I_fault_rms_kA`` (float)
    * ``X_over_R`` (float)

    Entradas com dados inconsistentes são puladas (com log de warning).
    """
    p = Path(path)
    with p.open("r", encoding="utf-8") as f:
        raw = json.load(f)

    if not isinstance(raw, list):
        raise ValueError(
            f"ct_system_data.json deve ser um array; got {type(raw).__name__}"
        )

    contexts: List[CTFaultContext] = []
    skipped = 0
    for i, entry in enumerate(raw):
        try:
            tag = str(entry.get("CT_TAG", f"CT-{i:03d}"))
            bus_in = str(entry.get("Bus_in") or tag)
            v_ll = float(entry.get("V_LL_kV") or 13.8)
            icc_kA = float(entry.get("I_fault_rms_kA") or 0.0)
            xr = float(entry.get("X_over_R") or 1.0)

            if icc_kA <= 0:
                skipped += 1
                continue

            ctx = CTFaultContext(
                bus_id=f"{bus_in} [{tag}]",
                rated_voltage_kV=v_ll,
                I_fault_rms_kA=icc_kA,
                X_over_R=max(xr, 0.1),
            )
            contexts.append(ctx)
        except (ValueError, TypeError, KeyError) as e:
            log.warning(
                "load_ct_system_data: pulando entry[%d] (%s): %s",
                i, entry.get("CT_TAG", "?"), e,
            )
            skipped += 1

    log.info(
        "load_ct_system_data: carregados %d TCs (%d pulados)",
        len(contexts), skipped,
    )
    return contexts


def list_cts_table(path: PathLike) -> List[dict]:
    """
    Sumário em formato tabular para popular QListWidget/QTableWidget
    no GUI. Cada item contém os campos chave do JSON original.

    Campos por linha:
    ``index``, ``CT_TAG``, ``Bus_in``, ``CT_Ratio``, ``V_LL_kV``,
    ``I_fault_rms_kA``, ``X_over_R``, ``Protection_Type``,
    ``Protected_Device``.
    """
    p = Path(path)
    with p.open("r", encoding="utf-8") as f:
        raw = json.load(f)

    if not isinstance(raw, list):
        raise ValueError(
            f"ct_system_data.json deve ser um array"
        )

    table = []
    for i, entry in enumerate(raw):
        table.append({
            "index": i,
            "CT_TAG": str(entry.get("CT_TAG", f"CT-{i:03d}")),
            "Bus_in": str(entry.get("Bus_in", "")),
            "CT_Ratio": str(entry.get("CT_Ratio", "")),
            "V_LL_kV": float(entry.get("V_LL_kV") or 0.0),
            "I_fault_rms_kA": float(entry.get("I_fault_rms_kA") or 0.0),
            "X_over_R": float(entry.get("X_over_R") or 0.0),
            "Protection_Type": str(entry.get("Protection_Type", "")),
            "Protected_Device": str(entry.get("Protected_Device", "")),
        })
    return table


# ---------------------------------------------------------------------------
# Sprint D helper — workflow refinamento
# ---------------------------------------------------------------------------


def next_ansi_class(current_C: float) -> float:
    """
    Próxima classe ANSI padrão acima do valor atual.

    ANSI C57.13 padrão: C100, C200, C400, C800.
    Acima de C800 → arredonda para próxima centena (custom).

    Examples
    --------
    >>> next_ansi_class(100.0)
    200.0
    >>> next_ansi_class(150.0)
    200.0
    >>> next_ansi_class(800.0)
    1600.0
    """
    standards = [100.0, 200.0, 400.0, 800.0]
    for s in standards:
        if s > current_C:
            return s
    # Acima do padrão — dobra para custom acima
    return math.ceil((current_C * 2) / 100.0) * 100.0


# ---------------------------------------------------------------------------
# Sprint F — Report writers (TXT + XLSX)
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# v1.4.5 — CTSelection bridge (PpComponent CT → CTSpec/CTFaultContext)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class CTSelection:
    """
    Bridge entre :class:`PpComponent` do canvas (type='CT') e os
    inputs do estudo de saturação (CTSpec + CTFaultContext).

    Razão (user gate v1.4.4): a análise de saturação SÓ deve rodar
    quando o usuário seleciona um TC no Modo Online do canvas e
    invoca a análise. Esse dataclass carrega os valores extraídos
    das properties do CT.ocomp.

    Os campos de **falta** (rated_voltage_kV, I_fault_rms_kA,
    X_over_R) vêm do contexto do barramento ao qual o TC está
    conectado — em v1.4.5 são preenchidos com defaults razoáveis
    para o usuário ajustar; em v1.5+ podem vir do study cache.
    """

    tag: str
    ratio_pri_A: float
    ratio_sec_A: float
    classe: str   # ANSI_C / IEC_P / IEC_PR
    c_rating_V: float
    rs_ohm: float
    rb_ohm: float
    kr: float
    t_relay_ms: float
    t_breaker_ms: float

    # Contexto de falta — defaults razoáveis (usuário ajusta)
    bus_id: str = "BUS-?"
    rated_voltage_kV: float = 13.8
    I_fault_rms_kA: float = 20.0
    X_over_R: float = 10.0

    @classmethod
    def from_pp_component(cls, comp) -> "CTSelection":
        """
        Constrói a partir de :class:`PpComponent` (type='CT').

        Lê as properties na **ordem do spec** (CT.ocomp tem 10
        properties: tag, ratio_pri, ratio_sec, classe, c_rating_V,
        rs_ohm, rb_ohm, kr, t_relay_ms, t_breaker_ms).

        Raises
        ------
        ValueError
            Se ``comp.type != 'CT'`` ou se o spec não estiver no
            registry.
        """
        if comp.type != "CT":
            raise ValueError(
                f"CTSelection requer componente do tipo 'CT' "
                f"(got {comp.type!r})"
            )

        from app.preprocessor.spec import get_default_registry
        spec = get_default_registry().get("CT")
        if spec is None:
            raise ValueError(
                "CT spec não está registrado em catalog_specs"
            )

        # Indexa properties pelo nome do spec
        values = {}
        for idx, p_spec in enumerate(spec.properties):
            if idx < len(comp.properties):
                values[p_spec.name] = comp.properties[idx].value

        def _f(key, default=0.0) -> float:
            try:
                return float(values.get(key, default) or default)
            except (TypeError, ValueError):
                return default

        def _s(key, default="") -> str:
            return str(values.get(key, default) or default)

        return cls(
            tag=_s("tag", "CT-?"),
            ratio_pri_A=_f("ratio_pri", 200.0),
            ratio_sec_A=_f("ratio_sec", 5.0),
            classe=_s("classe", "ANSI_C"),
            c_rating_V=_f("c_rating_V", 200.0),
            rs_ohm=_f("rs_ohm", 0.5),
            rb_ohm=_f("rb_ohm", 0.3),
            kr=_f("kr", 0.8),
            t_relay_ms=_f("t_relay_ms", 20.0),
            t_breaker_ms=_f("t_breaker_ms", 50.0),
        )


def _safe_filename(s: str) -> str:
    """Substitui chars não-alfanuméricos por '_' (paridade MATLAB regexprep)."""
    return re.sub(r"[^A-Za-z0-9]", "_", s)


def write_report_txt(
    report: CTSaturationReport, path: PathLike,
) -> None:
    """
    Gera relatório TXT (paridade ``generate_reports.m §TXT``).

    Estrutura:

    * Cabeçalho: TAG, data, RESULTADO FINAL
    * Seção 1: Especificação do TC
    * Seção 2: Dados do Sistema
    * Seção 3: Análise Dinâmica + dicas condicionais (Kr/PR)
    """
    p = Path(path)
    ct = report.ct
    fault = report.fault
    L1 = report.level1
    L2 = report.level2
    L3 = report.level3

    lines: List[str] = []
    lines.append("RELATÓRIO TÉCNICO — CT SATURATION STUDY (Olivas)")
    lines.append(f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"TAG: {ct.tag}")
    lines.append("-" * 60)
    lines.append(f"RESULTADO FINAL: {report.final_status}")
    lines.append(f"  → {report.final_detail}")
    lines.append("-" * 60)

    lines.append("")
    lines.append("1. ESPECIFICAÇÃO DO TC (selecionado):")
    lines.append(f"   Tipo/Norma:    {ct.tipo_TC}")
    lines.append(
        f"   Classe:        C{ct.CT_Rating_C:.0f}"
        + (f" ({ct.IEC_VA:.0f}VA)" if ct.IEC_VA else "")
    )
    lines.append(
        f"   Relação:       {ct.I_pri_nom_tap:.0f}/{ct.I_sn:.0f} "
        f"(full {ct.I_pri_nom_full:.0f})"
    )
    lines.append(
        f"   Remanência:    Kr = {ct.Kr:.2f} (considerada na simulação)"
    )
    lines.append(f"   R secundário:  Rs = {ct.R_s:.3f} Ω")
    lines.append(f"   Burden:        Rb = {ct.R_b:.3f} Ω (relé+cabo)")
    lines.append(
        f"   Z_loop total:  {ct.Z_loop:.3f} Ω"
    )

    lines.append("")
    lines.append("2. DADOS DO SISTEMA E CARGA:")
    lines.append(f"   Barra:         {fault.bus_id}")
    lines.append(f"   Tensão LL:     {fault.rated_voltage_kV:.2f} kV")
    lines.append(
        f"   Curto-Circuito: {fault.I_fault_rms_kA:.2f} kA "
        f"(I_peak {fault.I_fault_peak_A/1000:.1f} kA)"
    )
    lines.append(f"   X/R:           {fault.X_over_R:.2f}")
    lines.append(
        f"   Tempos:        Trip {ct.t_relay_ms:.0f} ms | "
        f"Extinção {ct.t_total_ms:.0f} ms"
    )

    lines.append("")
    lines.append("3. ANÁLISE DINÂMICA (3 níveis: ANSI estático / IEC joelho / Dinâmico RK4):")
    lines.append(
        f"   Nível 1 (ANSI):   V_tap = {L1.V_tap:.0f} V, "
        f"V_req = {L1.V_req:.0f} V → "
        f"{'PASSOU' if L1.passed else 'FALHOU'}"
    )
    lines.append(
        f"      Recomendação: {L1.recommended_class_C}"
    )
    if L2 is not None:
        lines.append(
            f"   Nível 2 (IEC):    Ek = {L2.Ek_actual:.0f} V, "
            f"Ek_req = {L2.Ek_req:.0f} V → "
            f"{'PASSOU' if L2.passed else 'FALHOU'}"
        )
        lines.append(f"      Recomendação: {L2.recommended_class}")
    lines.append(
        f"   Nível 3 (Dinâmico): t_sat = {L3.t_sat_ms:.1f} ms "
        f"(margem {L3.margin_ms:+.1f} ms) → {L3.status}"
    )
    lines.append(
        f"      Uso no Trip: {L3.sat_at_trip_pct:.1f}% | "
        f"Pico: {L3.sat_peak_pct:.1f}%"
    )

    # Dicas condicionais (paridade MATLAB)
    lines.append("")
    if L3.t_sat_ms < ct.t_relay_ms:
        lines.append(
            "⚠ CRÍTICO: O TC satura ANTES do relé atuar."
        )
        if ct.Kr > 0.1:
            lines.append(
                "  DICA: Considere classe IEC PR (low remanence) "
                "para ganhar margem dinâmica."
            )
        # Dica adicional Olivas: classe ANSI maior
        next_C = next_ansi_class(ct.CT_Rating_C)
        lines.append(
            f"  DICA: Considere classe ANSI maior — "
            f"C{ct.CT_Rating_C:.0f} → C{next_C:.0f}."
        )
    else:
        lines.append("✓ OK: O TC suporta o ciclo inicial de falta.")

    lines.append("")
    lines.append("-" * 60)
    lines.append("Citações normativas:")
    lines.append(f"  L1: {L1.citation}")
    if L2 is not None:
        lines.append(f"  L2: {L2.citation}")
    lines.append(f"  L3: {L3.citation}")

    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text("\n".join(lines), encoding="utf-8")
    log.info("write_report_txt: %s (%d linhas)", p, len(lines))


def write_report_xlsx(
    report: CTSaturationReport, path: PathLike,
) -> None:
    """
    Gera datasheet XLSX (paridade ``generate_reports.m §XLSX``).

    Tabela 2 colunas (Parâmetro/Valor) com 8 linhas chave + extras.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise ImportError(
            "openpyxl é necessário para exportar XLSX. "
            "Instale com: pip install openpyxl"
        ) from e

    p = Path(path)
    ct = report.ct
    fault = report.fault
    L3 = report.level3

    wb = Workbook()
    ws = wb.active
    ws.title = "CT Datasheet"

    # Header
    ws["A1"] = "Parâmetro"
    ws["B1"] = "Valor"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="143250")
    for cell in (ws["A1"], ws["B1"]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 8 linhas core (paridade MATLAB)
    rows = [
        ("TAG", ct.tag),
        ("Relação",
         f"{ct.I_pri_nom_tap:.0f}/{ct.I_sn:.0f}"),
        ("Classe",
         f"C{ct.CT_Rating_C:.0f}"
         + (f" ({ct.IEC_VA:.0f}VA)" if ct.IEC_VA else "")),
        ("Tipo/Norma", ct.tipo_TC),
        ("Kr (remanência)", f"{ct.Kr:.2f}"),
        ("Icc (A, RMS)", f"{fault.I_fault_rms_A:.0f}"),
        ("t_sat (ms)",
         f"{L3.t_sat_ms:.1f}" if L3.t_sat_ms != float("inf") else "∞"),
        ("Resultado", report.final_status),
    ]
    # Linhas extras Olivas (não em MATLAB mas úteis)
    rows.extend([
        ("Bus_in", fault.bus_id),
        ("X/R", f"{fault.X_over_R:.2f}"),
        ("V_LL (kV)", f"{fault.rated_voltage_kV:.2f}"),
        ("Z_loop (Ω)", f"{ct.Z_loop:.4f}"),
        ("Margem (ms)", f"{L3.margin_ms:+.1f}"),
        ("Uso no Trip (%)", f"{L3.sat_at_trip_pct:.1f}"),
        ("Pico Saturação (%)", f"{L3.sat_peak_pct:.1f}"),
        ("Detalhe", report.final_detail),
    ])

    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    # Auto-largura razoável
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    log.info("write_report_xlsx: %s (%d linhas)", p, len(rows) + 1)


# ---------------------------------------------------------------------------
# Convenience — gera ambos os formatos com nome padrão paridade MATLAB
# ---------------------------------------------------------------------------


def write_reports_default_names(
    report: CTSaturationReport, output_dir: PathLike,
) -> tuple[Path, Path]:
    """
    Gera ``Relatorio_<TAG>.txt`` e ``Datasheet_<TAG>.xlsx`` num
    diretório (paridade ``generate_reports.m`` linha 4-5).

    Returns
    -------
    (txt_path, xlsx_path)
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    tag_safe = _safe_filename(report.ct.tag)
    txt_path = out / f"Relatorio_{tag_safe}.txt"
    xlsx_path = out / f"Datasheet_{tag_safe}.xlsx"
    write_report_txt(report, txt_path)
    try:
        write_report_xlsx(report, xlsx_path)
    except ImportError as e:
        log.warning("XLSX skipped: %s", e)
        return txt_path, txt_path
    return txt_path, xlsx_path
