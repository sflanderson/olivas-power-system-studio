"""
app.postprocessor.dapper_io — TXT/XLSX exporters para
Demand Load Study (v1.5.1 Sprint E).

Paridade com PTW DAPPER ``LOAD SCHEDULE FOR <BUS>`` report e
com lessons learned das versões CT-SAT (TXT estruturado em
seções + XLSX com header colorido).

API
====

* :func:`write_demand_report_txt(report, path)` — relatório TXT
* :func:`write_demand_report_xlsx(report, path)` — datasheet XLSX
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Union

from app.core.logging_config import get_logger
from app.postprocessor.studies.demand_load import DemandReport

log = get_logger(__name__)

PathLike = Union[str, Path]


# ---------------------------------------------------------------------------
# TXT export
# ---------------------------------------------------------------------------


def write_demand_report_txt(
    report: DemandReport, path: PathLike,
) -> None:
    """
    Escreve relatório TXT (paridade PTW DAPPER LOAD SCHEDULE).

    Estrutura:

    * Cabeçalho com timestamp + título
    * Resumo total (Connected/Demand/Design + DF + PF)
    * Por categoria (tabela com Connected/Demand/Design)
    * Detalhe por entry
    * Citações normativas (NEC 220.x + NBR 5410)
    """
    p = Path(path)
    lines = []

    lines.append("=" * 78)
    lines.append("  RELATÓRIO TÉCNICO — DEMAND LOAD STUDY (DAPPER)")
    lines.append(
        "  Olivas Power System Studio  ·  "
        f"Gerado em {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    lines.append("=" * 78)
    lines.append("")

    # Summary já é texto formatado em report.summary()
    lines.append(report.summary())
    lines.append("")

    # Citações normativas
    lines.append("REFERÊNCIAS NORMATIVAS")
    lines.append("-" * 78)
    nec_sections = sorted({
        cr.nec_section for cr in report.by_category
        if cr.nec_section
    })
    if nec_sections:
        lines.append(
            "NEC 2023 sections aplicadas:"
        )
        for s in nec_sections:
            lines.append(f"  · NEC §{s}")
    lines.append(
        "Outras: NBR 5410:2004 §6 (Cálculo de Cargas), "
        "PTW DAPPER (Reference §1.2-1.4)"
    )
    lines.append("")
    lines.append("=" * 78)

    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text("\n".join(lines), encoding="utf-8")
    log.info(
        "write_demand_report_txt: %s (%d entries, %d categorias)",
        p, len(report.entries), len(report.by_category),
    )


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


def write_demand_report_xlsx(
    report: DemandReport, path: PathLike,
) -> None:
    """
    Escreve datasheet XLSX com 2 sheets:

    * **Resumo** — Connected/Demand/Design por categoria + total
    * **Detalhe** — entries originais (item-by-item)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise ImportError(
            "openpyxl é necessário para XLSX. "
            "Instale com: pip install openpyxl"
        ) from e

    p = Path(path)
    wb = Workbook()

    # ---- Sheet 1: Resumo por categoria ----
    ws = wb.active
    ws.title = "Resumo"
    headers = [
        "Categoria", "NEC §", "Tipo", "N Entries",
        "Connected (kVA)", "Demand (kVA)",
        "Design (kVA)", "PF", "LCL",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="143250")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for cr in report.by_category:
        ws.cell(row=row, column=1, value=cr.category_label_pt)
        ws.cell(row=row, column=2, value=cr.nec_section or "—")
        # Buscamos load_type via catalog (não está em CategoryResult)
        from app.postprocessor.dapper_catalog import get_category
        cat = get_category(cr.category_code)
        ws.cell(
            row=row, column=3,
            value=cat.load_type.value if cat else "—",
        )
        ws.cell(row=row, column=4, value=cr.n_entries)
        ws.cell(row=row, column=5, value=round(cr.connected_kVA, 1))
        ws.cell(row=row, column=6, value=round(cr.demand_kVA, 1))
        ws.cell(row=row, column=7, value=round(cr.design_kVA, 1))
        ws.cell(row=row, column=8, value=round(cr.pf, 3))
        ws.cell(row=row, column=9, value=round(cr.lcl_factor, 2))
        row += 1

    # Linha de totais
    total_row = row
    bold = Font(bold=True)
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws.cell(
        row=total_row, column=5,
        value=round(report.total_connected_kVA, 1),
    ).font = bold
    ws.cell(
        row=total_row, column=6,
        value=round(report.total_demand_kVA, 1),
    ).font = bold
    ws.cell(
        row=total_row, column=7,
        value=round(report.total_design_kVA, 1),
    ).font = bold
    ws.cell(
        row=total_row, column=8,
        value=round(report.weighted_pf, 3),
    ).font = bold

    # Auto-width
    widths = {1: 38, 2: 8, 3: 6, 4: 9, 5: 16, 6: 13, 7: 13, 8: 7, 9: 6}
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    # ---- Sheet 2: Detalhe (entries) ----
    ws2 = wb.create_sheet("Detalhe")
    detail_headers = [
        "Nome", "Categoria", "Connected (kVA)", "PF override",
    ]
    for col, h in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, e in enumerate(report.entries, start=2):
        ws2.cell(row=i, column=1, value=e.name)
        ws2.cell(row=i, column=2, value=e.category_code)
        ws2.cell(row=i, column=3, value=round(e.connected_kVA, 1))
        ws2.cell(
            row=i, column=4,
            value="" if e.pf is None else round(e.pf, 3),
        )

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 14

    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    log.info(
        "write_demand_report_xlsx: %s (%d categorias, %d entries)",
        p, len(report.by_category), len(report.entries),
    )
